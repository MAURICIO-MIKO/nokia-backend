import os
import ipaddress
import pandas as pd
import openpyxl as xl

def convertir_xls_a_xlsx(xls_path):
    """
    Convierte cualquier archivo .xls a .xlsx aunque xlrd no lo soporte.
    Se apoya en openpyxl reconstruyendo la estructura.
    Funciona en Render.
    """
    import xlrd
    import openpyxl

    libro_xls = xlrd.open_workbook(xls_path)
    libro_xlsx = openpyxl.Workbook()
    libro_xlsx.remove(libro_xlsx.active)

    for nombre_hoja in libro_xls.sheet_names():
        sh = libro_xls.sheet_by_name(nombre_hoja)
        hoja_nueva = libro_xlsx.create_sheet(title=nombre_hoja)

        for row in range(sh.nrows):
            for col in range(sh.ncols):
                hoja_nueva.cell(row=row + 1, col=col + 1).value = sh.cell_value(row, col)

    nuevo = xls_path + ".xlsx"
    libro_xlsx.save(nuevo)
    return nuevo


def main(xls, xml):
    """Procesa un Excel y plantilla XML (4G + 5G), genera un archivo XML en /tmp/salidas y devuelve su ruta."""

    # === CONVERTIR XLS A XLSX (VERSIÓN COMPATIBLE CON RENDER) ===
    xlsx = convertir_xls_a_xlsx(xls)

    dic = {}
    wb = xl.load_workbook(xlsx)

    # === RECOGE VARIABLES DE 5G ===
    ws = wb["5G"]
    dic["##site##"] = f"{buscaCelda(ws, 'Site NAME').replace(' ', '_')}_{buscaCelda(ws, 'Cod. CelSig')}"
    dic["##mr_bts_name##"] = buscaCelda(ws, "gNodeB NAME")[3:]
    dic["##mr_bts_id##"] = buscaCelda(ws, "mrBTSId")
    dic["##gnb_id##"] = buscaCelda(ws, "gNodeB id")
    dic["##ip_gnb_oym##"] = buscaCelda(ws, "O&M gNB IP")
    dic["##mask_gnb_oym##"] = ipaddress.IPv4Network(f"0.0.0.0/{buscaCelda(ws, 'O&M NetMask')}").prefixlen
    dic["##gateway_gnb_oym##"] = buscaCelda(ws, "O&M Gateway")
    dic["##vlan_gnb_oym##"] = buscaCelda(ws, "O&M VLAN")
    dic["##ip_ioms_prim_oym##"] = buscaCelda(ws, "O&M iOMS-prim IP")
    dic["##ip_ioms_sec_oym##"] = buscaCelda(ws, "O&M iOMS-sec IP")
    dic["##ip_netact_oym##"] = buscaCelda(ws, "O&M NetAct Subnet").split("/")[0]
    dic["##mask_netact_oym##"] = buscaCelda(ws, "O&M NetAct Subnet").split("/")[1]
    dic["##ip_pki_oym##"] = buscaCelda(ws, "O&M PKI Server IP")
    dic["##ip_gnb_sync##"] = buscaCelda(ws, "Sync gNB IP")
    dic["##mask_gnb_sync##"] = ipaddress.IPv4Network(f"0.0.0.0/{buscaCelda(ws, 'Sync NetMask')}").prefixlen
    dic["##gateway_sync##"] = buscaCelda(ws, "Sync Gateway")
    dic["##vlan_sync##"] = buscaCelda(ws, "Sync VLAN")
    dic["##ip_top1_sync##"] = buscaCelda(ws, "Sync ToP1 IP")
    dic["##ip_top2_sync##"] = buscaCelda(ws, "Sync ToP2 IP")
    dic["##ip_gnb_s1_outer##"] = buscaCelda(ws, "S1 gNB IP/OUTER IP")
    dic["##mask_s1_x2##"] = ipaddress.IPv4Network(f"0.0.0.0/{buscaCelda(ws, 'S1-X2 NetMask')}").prefixlen
    dic["##gateway_s1_x2##"] = buscaCelda(ws, "S1-X2 Gateway")
    dic["##vlan_s1_x2##"] = buscaCelda(ws, "S1-X2 VLAN")
    dic["##ip_gnb_inner##"] = buscaCelda(ws, "INNER IP")
    dic["##gateway_security##"] = buscaCelda(ws, "Security Gateway IP")
    dic["##tac##"] = buscaCelda(ws, "TAC")

    # === AMF CONFIGURACIÓN ===
    dic["##AMF_mde1##"] = buscaCelda(ws, "AMF nvpcc-amf01/2/3-mde1").split("/")[0]
    dic["##AMF_mno1##"] = buscaCelda(ws, "AMF nvpcc-amf01/2/3-mno1").split("/")[0]
    dic["##AMF_bep1##"] = buscaCelda(ws, "AMF nvpcc-amf01/2/3-bep1").split("/")[0]
    dic["##AMF_btb1##"] = buscaCelda(ws, "AMF nvpcc-amf01/2/3-btb1").split("/")[0]
    dic["##AMF_mad1##"] = buscaCelda(ws, "AMF nvpcc-amf01/2/3-mad1").split("/")[0]

    # === PARÁMETROS RADIO NR3500 ===
    dic["##Zero correlation zone config NR3500##"] = buscaCelda(ws, "Zero correlation zone config NR3500")
    dic["##Prach Configuration Index NR3500##"] = buscaCelda(ws, "Prach Configuration Index NR3500")
    dic["##Phys Cell Id NR3500##"] = buscaCelda(ws, "Phys Cell Id NR3500")
    dic["##PRACH root sequence Index NR3500##"] = buscaCelda(ws, "PRACH root sequence Index NR3500")

    # === ADVERTENCIA SI HAY VARIABLES VACÍAS ===
    for v, val in dic.items():
        if not val:
            print(f"⚠️ WARNING: Variable {v} vacía")

    # === SUSTITUIR VARIABLES EN XML ===
    with open(xml, "r", encoding="utf-8") as archivo:
        contenido = archivo.read()

    for v, valor in dic.items():
        contenido = contenido.replace(v, str(valor))

    enb = f"ENB{dic['##mr_bts_name##']}".replace("_", "-")
    contenido = contenido.replace(f"ENB{dic['##mr_bts_name##']}", enb)

    # === CREAR CARPETA /tmp/salidas (para Render) ===
    output_dir = os.path.join("/tmp", "salidas")
    os.makedirs(output_dir, exist_ok=True)

    # === GUARDAR ARCHIVO FINAL CON NOMBRE AUTOMÁTICO ===
    nombre_salida = f"_{dic['##site##']}.xml"
    ruta_salida = os.path.join(output_dir, nombre_salida)

    with open(ruta_salida, "w", encoding="utf-8") as f:
        f.write(contenido)

    print(f"✅ Archivo {ruta_salida} generado correctamente.")
    os.remove(xlsx)
    return ruta_salida  # ⚠️ Devuelve la ruta completa (Render lo necesita)


def buscaCelda(hoja, valor_buscado):
    """Busca un valor en la primera fila y devuelve el valor de la segunda."""
    for celda in hoja[1]:
        if celda.value == valor_buscado:
            celda_objetivo = f"{celda.column_letter}2"
            return hoja[celda_objetivo].value
    return ""
